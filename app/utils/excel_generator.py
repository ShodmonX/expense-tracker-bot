import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import Dict, List

async def generate_excel_report(report_data: Dict, filename: str = None) -> str:
    """Generate Excel report from report data"""
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment(horizontal="center")
    
    currency_font = Font(size=11)
    currency_alignment = Alignment(horizontal="right")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:E1')
    ws['A1'] = report_data.get("period", "Hisobot")
    ws['A1'].font = title_font
    ws['A1'].alignment = title_alignment
    
    # Headers
    headers = ["Sana", "Turi", "Kategoriya", "Miqdor (so'm)", "Izoh"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data
    row_num = 4
    
    # Add income data first
    incomes = report_data.get("incomes", [])
    for income in incomes:
        ws.cell(row=row_num, column=1, value=income.date.strftime("%d.%m.%Y")).border = thin_border
        ws.cell(row=row_num, column=2, value="KIRIM").border = thin_border
        ws.cell(row=row_num, column=3, value=income.category).border = thin_border
        
        amount_cell = ws.cell(row=row_num, column=4, value=income.amount)
        amount_cell.number_format = '#,##0'
        amount_cell.font = currency_font
        amount_cell.alignment = currency_alignment
        amount_cell.border = thin_border
        
        ws.cell(row=row_num, column=5, value=income.description or "").border = thin_border
        row_num += 1
    
    # Add expense data
    expenses = report_data.get("expenses", [])
    for expense in expenses:
        ws.cell(row=row_num, column=1, value=expense.date.strftime("%d.%m.%Y")).border = thin_border
        ws.cell(row=row_num, column=2, value="XARAJAT").border = thin_border
        ws.cell(row=row_num, column=3, value=expense.category).border = thin_border
        
        amount_cell = ws.cell(row=row_num, column=4, value=expense.amount)
        amount_cell.number_format = '#,##0'
        amount_cell.font = currency_font
        amount_cell.alignment = currency_alignment
        amount_cell.border = thin_border
        
        ws.cell(row=row_num, column=5, value=expense.description or "").border = thin_border
        row_num += 1
    
    # Summary
    row_num += 1
    ws.cell(row=row_num, column=1, value="JAMI KIRIM:").font = Font(bold=True)
    income_total_cell = ws.cell(row=row_num, column=3, value=report_data.get("total_income", 0))
    income_total_cell.font = Font(bold=True)
    income_total_cell.number_format = '#,##0'
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="JAMI XARAJAT:").font = Font(bold=True)
    expense_total_cell = ws.cell(row=row_num, column=3, value=report_data.get("total_expenses", 0))
    expense_total_cell.font = Font(bold=True)
    expense_total_cell.number_format = '#,##0'
    
    row_num += 1
    ws.cell(row=row_num, column=1, value="BALANS:").font = Font(bold=True)
    balance = report_data.get("balance", 0)
    balance_total_cell = ws.cell(row=row_num, column=3, value=balance)
    balance_total_cell.font = Font(bold=True)
    balance_total_cell.number_format = '#,##0'
    if balance < 0:
        balance_total_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Category summary
    if report_data.get("category_totals"):
        row_num += 2
        ws.cell(row=row_num, column=1, value="Kategoriyalar bo'yicha:").font = Font(bold=True)
        row_num += 1
        
        category_totals = report_data.get("category_totals", {})
        for category, amount in sorted(category_totals.items(), key=lambda x: x[1], reverse=True):
            ws.cell(row=row_num, column=1, value=category)
            amount_cell = ws.cell(row=row_num, column=2, value=amount)
            amount_cell.number_format = '#,##0'
            row_num += 1
    
    # Monthly summary for yearly report
    if report_data.get("monthly_totals"):
        row_num += 2
        ws.cell(row=row_num, column=1, value="Oylik hisobot:").font = Font(bold=True)
        row_num += 1
        
        monthly_totals = report_data.get("monthly_totals", {})
        for month, data in monthly_totals.items():
            month_name = datetime(2024, month, 1).strftime("%B")
            ws.cell(row=row_num, column=1, value=month_name)
            
            if isinstance(data, dict):
                # New format - show balance
                balance = data.get('balance', 0)
                amount_cell = ws.cell(row=row_num, column=2, value=balance)
                amount_cell.number_format = '#,##0'
                if balance < 0:
                    amount_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            else:
                # Old format - show amount
                amount_cell = ws.cell(row=row_num, column=2, value=data)
                amount_cell.number_format = '#,##0'
            
            row_num += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename if not provided
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reports/hisobot_{timestamp}.xlsx"
    
    # Ensure reports directory exists
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    
    # Save workbook
    wb.save(filename)
    
    return filename
